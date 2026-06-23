import pandas as pd
from openpyxl import load_workbook
import re
import sys

def is_red_font(color_obj):
    """Check if font color is red."""
    if not color_obj:
        return False
    if hasattr(color_obj, 'rgb') and color_obj.rgb:
        rgb = str(color_obj.rgb).upper()
        if rgb == 'FFFF0000' or rgb == 'FF0000':
            return True
    if hasattr(color_obj, 'indexed') and color_obj.indexed == 2:
        return True
    return False

def get_red_text_indices(file_path, sheet_name=None):
    """Get indices of rows with red text."""
    wb = load_workbook(file_path, data_only=False)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    red_indices = set()

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row_idx, col_idx)
            if cell.font and cell.font.color and is_red_font(cell.font.color):
                red_indices.add(row_idx - 2)
                break
    return red_indices

def fix_capitalization(name):
    """Fix all-caps or improper capitalization."""
    if pd.isna(name) or name == '':
        return name
    name = str(name).strip()

    # If the name is all uppercase or all lowercase, convert to title case
    if name.isupper() or name.islower():
        return name.title()

    # Otherwise return as-is (mixed case is probably intentional)
    return name

def parse_name(full_name, default_last_name=''):
    """Parse a full name into first and last name."""
    if pd.isna(full_name) or full_name == '':
        return None, None, False
    full_name = str(full_name).strip()

    # Remove common suffixes/notes that aren't part of the name
    # Examples: "John Smith - son", "Jane Doe (daughter)", "Bob Jones, grandson"
    full_name = re.split(r'\s*[-–—]\s*(?:son|daughter|grandson|granddaughter|parent|mother|father|sister|brother|friend|guest)\s*$', full_name, flags=re.IGNORECASE)[0]
    full_name = re.sub(r'\s*\([^)]*(?:son|daughter|grandson|granddaughter|parent|mother|father|sister|brother|friend|guest)[^)]*\)\s*$', '', full_name, flags=re.IGNORECASE)
    full_name = full_name.strip()

    parts = full_name.split()
    if len(parts) == 0:
        return None, None, False
    elif len(parts) == 1:
        # Only first name, using default
        return fix_capitalization(parts[0]), fix_capitalization(default_last_name) if default_last_name else default_last_name, True
    else:
        first_name = ' '.join([fix_capitalization(p) for p in parts[:-1]])
        last_name = fix_capitalization(parts[-1])
        return first_name, last_name, False

def parse_guests_string(guests_str):
    """Parse a comma-separated string of guest names."""
    if pd.isna(guests_str) or guests_str == '':
        return []
    guests_str = str(guests_str)
    guests_str = re.sub(r'\s+and\s+', ', ', guests_str)
    names = [name.strip() for name in guests_str.split(',')]
    return [name for name in names if name]

def extract_names_from_message(message_str):
    """Extract names from message/notes field."""
    if pd.isna(message_str) or message_str == '':
        return []

    message_str = str(message_str)
    name_pattern = r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,2})\b'
    potential_names = re.findall(name_pattern, message_str)

    exclude_words = ['Please', 'Thank', 'Could', 'Would', 'Will', 'The', 'And', 'But', 'For', 'From',
                     'With', 'This', 'That', 'Have', 'Been', 'Were', 'Their', 'There', 'What', 'When']

    names = []
    for name in potential_names:
        if not any(name.startswith(word) for word in exclude_words):
            parts = name.split()
            if len(parts) >= 2:
                names.append(name)

    return names

all_records = []

# Process MASTER file
print("Processing MASTER file...")
sheet_name = 'CH&MH, Culbreth, JPJ. Sat & Sun'
red_indices = get_red_text_indices('C:/Users/Ben/Downloads/Finals Weekend Parking List (MASTER) 2026.xlsx', sheet_name)
master = pd.read_excel('C:/Users/Ben/Downloads/Finals Weekend Parking List (MASTER) 2026.xlsx', sheet_name=sheet_name)

for idx, row in master.iterrows():
    if idx in red_indices:
        continue

    first_name = fix_capitalization(row.get('First_Name', ''))
    last_name = fix_capitalization(row.get('Last_Name', ''))
    affiliation_raw = row.get('Affliation', '')  # Note: misspelled in source
    day = row.get('Day', '')

    # For Culbreth parking, check the "Culbreth Garage / JPJ Garage" column
    try:
        culbreth_passes = float(row.get('Culbreth Garage / JPJ Garage', 0)) if pd.notna(row.get('Culbreth Garage / JPJ Garage')) else 0
    except:
        culbreth_passes = 0

    # Skip if no Culbreth parking needed
    if culbreth_passes <= 0:
        continue

    # Determine ceremony day
    day_str = str(day).strip().lower()
    if day_str == 'both':
        ceremony_day = 'Both'
    elif 'sat' in day_str:
        ceremony_day = 'Saturday'
    elif 'sun' in day_str:
        ceremony_day = 'Sunday'
    else:
        ceremony_day = 'Both'  # Default if unclear

    # Use the actual affiliation from the file, default to "Special Guest" if empty
    affiliation = affiliation_raw if pd.notna(affiliation_raw) and affiliation_raw != '' else 'Special Guest'

    all_records.append({
        'First Name': first_name,
        'Last Name': last_name,
        'Affiliation': affiliation,
        'Ceremony Day': ceremony_day,
        'Number of Passes': int(culbreth_passes),
        'Notes': ''
    })

# Process ceremony files
ceremonies = [
    ('C:/Users/Ben/Downloads/Saturday (2)_4.28.2026.xlsx', 'Saturday'),
    ('C:/Users/Ben/Downloads/Saturday (4)_4.28.2026.xlsx', 'Saturday'),
    ('C:/Users/Ben/Downloads/Sunday (2)_4.28.2026.xlsx', 'Sunday'),
    ('C:/Users/Ben/Downloads/Sunday (4)_4.28.2026.xlsx', 'Sunday')
]

# Process Faculty Marshals file
print("\nProcessing Faculty Marshals...")
faculty_file = 'C:/Users/Ben/Downloads/Faculty Marshals 2026_List for Parking.xlsx'
faculty_red_indices = get_red_text_indices(faculty_file)
faculty_df = pd.read_excel(faculty_file)
print(f"  Total rows: {len(faculty_df)}, Excluding red: {len(faculty_red_indices)}, Keeping: {len(faculty_df) - len(faculty_red_indices)}")

for idx, row in faculty_df.iterrows():
    if idx in faculty_red_indices:
        continue

    first_name = fix_capitalization(row.get('First Name', ''))
    last_name = fix_capitalization(row.get('Last Name', ''))

    if pd.isna(last_name) or last_name == '':
        continue

    # Add single record with "Both" for ceremony day
    all_records.append({
        'First Name': first_name if pd.notna(first_name) else '',
        'Last Name': last_name if pd.notna(last_name) else '',
        'Affiliation': 'Faculty Marshal',
        'Ceremony Day': 'Both',
        'Number of Passes': 1,
        'Notes': ''
    })

for file_path, ceremony_time in ceremonies:
    print(f"Processing {ceremony_time}...")
    red_indices = get_red_text_indices(file_path)
    df = pd.read_excel(file_path)

    for idx, row in df.iterrows():
        if idx in red_indices:
            continue

        family_last_name = fix_capitalization(row.get('Last Name', ''))
        first_name = fix_capitalization(row.get('First Name', ''))
        guests_str = row.get('Guests', '')
        grad_col = "Please enter the graduating students' name:"
        grad_student = row.get(grad_col, '')
        message_str = row.get('Message to the Event Coordinator', '')
        total_tickets = row.get('Total Ticket Request', 0)

        if pd.isna(family_last_name) or family_last_name == '':
            continue

        # Skip summary/total rows
        if str(family_last_name).lower().strip() in ['total', 'subtotal', 'grand total']:
            continue

        try:
            total_tickets = int(float(total_tickets)) if pd.notna(total_tickets) else 0
        except:
            total_tickets = 0

        # Add main contact
        all_records.append({
            'First Name': first_name if pd.notna(first_name) else '',
            'Last Name': family_last_name if pd.notna(family_last_name) else '',
            'Affiliation': 'Special Guest',
            'Ceremony Day': ceremony_time,
            'Number of Passes': total_tickets,
            'Notes': ''
        })

        # Add graduating student
        if pd.notna(grad_student) and grad_student != '':
            grad_first, grad_last, used_default = parse_name(grad_student, default_last_name=family_last_name)
            if grad_first:
                note = 'Last name inferred from family' if used_default else ''
                all_records.append({
                    'First Name': grad_first,
                    'Last Name': grad_last if grad_last else family_last_name,
                    'Affiliation': 'Special Guest',
                    'Ceremony Day': ceremony_time,
                    'Number of Passes': 0,
                    'Notes': note
                })

        # Add guests from Guests column
        guest_names = parse_guests_string(guests_str)
        for guest_name in guest_names:
            guest_first, guest_last, used_default = parse_name(guest_name, default_last_name=family_last_name)
            if guest_first:
                note = 'Last name inferred from family' if used_default else ''
                all_records.append({
                    'First Name': guest_first,
                    'Last Name': guest_last if guest_last else family_last_name,
                    'Affiliation': 'Special Guest',
                    'Ceremony Day': ceremony_time,
                    'Number of Passes': 0,
                    'Notes': note
                })

        # Extract names from message field
        message_names = extract_names_from_message(message_str)
        for message_name in message_names:
            msg_first, msg_last, used_default = parse_name(message_name, default_last_name=family_last_name)
            if msg_first:
                note = 'From message field'
                if used_default:
                    note += '; Last name inferred from family'
                all_records.append({
                    'First Name': msg_first,
                    'Last Name': msg_last if msg_last else family_last_name,
                    'Affiliation': 'Special Guest',
                    'Ceremony Day': ceremony_time,
                    'Number of Passes': 0,
                    'Notes': note
                })

df_final = pd.DataFrame(all_records)

# Reorder columns: Last Name first, then First Name
df_final = df_final[['Last Name', 'First Name', 'Affiliation', 'Ceremony Day', 'Number of Passes', 'Notes']]

# Handle duplicates: For people appearing multiple times, keep the record with the highest pass count
# and prefer non-"Special Guest" affiliations (staff categories)
print(f"\nTotal records before deduplication: {len(df_final)}")

# Create a priority score for affiliations (non-Special Guest gets higher priority)
def affiliation_priority(affiliation):
    if affiliation == 'Special Guest':
        return 0
    else:
        return 1

df_final['affiliation_priority'] = df_final['Affiliation'].apply(affiliation_priority)

# Sort by name, then priority (descending), then pass count (descending)
df_final = df_final.sort_values(
    by=['Last Name', 'First Name', 'affiliation_priority', 'Number of Passes'],
    ascending=[True, True, False, False]
)

# Keep only the first (best) record for each person
df_final = df_final.drop_duplicates(subset=['Last Name', 'First Name'], keep='first')

# Drop the temporary priority column
df_final = df_final.drop(columns=['affiliation_priority'])

print(f"Total records after deduplication: {len(df_final)}")

# Sort by Last Name, then First Name
df_final = df_final.sort_values(by=['Last Name', 'First Name'], ascending=[True, True])

# Save
output_file = 'C:/Users/Ben/Downloads/Graduation_Parking_Reconciled_2026_v2.xlsx'
df_final.to_excel(output_file, index=False)

print(f"\nTotal records: {len(df_final)}")

# Check Barr records
barr_records = df_final[df_final['Last Name'].str.contains('Barr', case=False, na=False)]
print(f"\nBARR FAMILY RECORDS:")
print(barr_records.to_string(index=False))
